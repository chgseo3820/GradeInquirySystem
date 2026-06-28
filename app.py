# -*- coding: utf-8 -*-
"""
ScoreQuery — 성적 조회 시스템 백엔드 (Flask 개발/로컬 운영용)

학생 본인 성적 조회 API와, 운영자가 GitHub Pages 배포 데이터를
암호화 저장(`docs/data.enc.json`)할 수 있도록 도와주는 관리자 API를 제공합니다.

⚠️ 보안 주의
  - 본 서버는 **로컬 운영자 PC에서만 동작**시키는 것을 전제로 합니다.
    외부 네트워크에 노출하지 마십시오. 기본 바인딩은 `127.0.0.1` 입니다.
  - `/api/save_*` 관리자 엔드포인트는 `SCOREQUERY_ADMIN_TOKEN` 환경변수에
    설정된 토큰을 `X-Admin-Token` 헤더 또는 요청 본문에 함께 보낼 때만
    동작합니다. (관리자 PC에서 자기 자신만 호출하므로 정적 토큰으로 충분)
  - `/api/score`는 단순 IP 기반 in-memory rate limiting을 적용해
    학번 무차별 대입을 완화합니다.
"""

import os
import re
import json
import time
import hmac
import hashlib
import secrets
import threading
from collections import defaultdict, deque
from datetime import datetime, timezone

from flask import Flask, render_template, request, jsonify, send_from_directory, redirect, session
import openpyxl
from werkzeug.middleware.proxy_fix import ProxyFix
from werkzeug.security import check_password_hash, generate_password_hash

try:
    from argon2 import PasswordHasher
    from argon2 import exceptions as argon2_exceptions
except Exception:  # pragma: no cover - optional production dependency
    PasswordHasher = None
    argon2_exceptions = None

try:
    import bcrypt
except Exception:  # pragma: no cover - optional production dependency
    bcrypt = None

from scorequery_crypto import (
    PASSPHRASE_ENV,
    EncryptionConfigError,
    get_env_passphrase,
    write_encrypted_json,
)


# ──────────────────────────────────────────────
# 설정
# ──────────────────────────────────────────────
ADMIN_TOKEN_ENV = "SCOREQUERY_ADMIN_TOKEN"
SECRET_KEY_ENV = "SCOREQUERY_SECRET_KEY"

# Flask 기본 바인딩 (로컬 운영자 PC 전용). 외부 노출이 필요하면 환경변수로 override.
DEFAULT_HOST = os.environ.get("SCOREQUERY_HOST", "127.0.0.1")
DEFAULT_PORT = int(os.environ.get("SCOREQUERY_PORT", "5000"))

# CORS 허용 출처 (콤마 구분). 빈 값/미설정 시 동일 출처(=헤더 미부착)만 허용.
ALLOWED_ORIGINS = [
    o.strip() for o in os.environ.get("SCOREQUERY_ALLOWED_ORIGINS", "").split(",") if o.strip()
]

# 학번 입력 길이 상한 (Excel 학번이 8~10자리이므로 보수적으로 12까지만 허용)
MAX_STUDENT_ID_LEN = 12

# 단순 in-memory rate limit: 60초 윈도우에 IP당 최대 N회 시도
RATE_WINDOW_SEC = 60
RATE_MAX_PER_WINDOW = 30
_rate_buckets: dict[str, deque] = defaultdict(deque)


CONFIG_FILE = os.path.join("config.json")
AUTH_DB_FILE = os.environ.get("SCOREQUERY_AUTH_DB", os.path.join("instance", "professors.json"))
VIEW_LOG_FILE = os.environ.get("SCOREQUERY_VIEW_LOG_FILE", os.path.join("instance", "view_logs.jsonl"))
PASSWORD_HASH_ALGO = os.environ.get("SCOREQUERY_PASSWORD_HASH", "argon2").strip().lower()


def _bool_env(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None:
        return default
    return raw.strip().lower() in {"1", "true", "yes", "on"}


REQUIRE_HTTPS = _bool_env("SCOREQUERY_REQUIRE_HTTPS")
TRUST_PROXY = _bool_env("SCOREQUERY_TRUST_PROXY")
SESSION_COOKIE_SECURE = _bool_env("SCOREQUERY_SESSION_COOKIE_SECURE", REQUIRE_HTTPS)
SESSION_COOKIE_SAMESITE = os.environ.get(
    "SCOREQUERY_SESSION_COOKIE_SAMESITE",
    "None" if ALLOWED_ORIGINS and SESSION_COOKIE_SECURE else "Lax",
)


def _get_configured_access_code() -> str:
    """설정 파일(config.json) 또는 환경변수에서 6자리 접속 비밀번호 조회"""
    env_val = os.environ.get("SCOREQUERY_ACCESS_CODE")
    if env_val:
        return env_val.strip()

    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                return str(cfg.get("access_code", "")).strip()
        except Exception:
            pass
    return ""


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _read_json_file(path: str, default):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def _write_json_file(path: str, data) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    tmp_path = f"{path}.tmp"
    with open(tmp_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    os.replace(tmp_path, path)


def _load_auth_users() -> list[dict]:
    data = _read_json_file(AUTH_DB_FILE, [])
    return data if isinstance(data, list) else []


def _save_auth_users(users: list[dict]) -> None:
    _write_json_file(AUTH_DB_FILE, users)


def _sanitize_auth_user(user: dict | None) -> dict | None:
    if not user:
        return None
    allowed = ("email", "name", "univ", "dept", "phone", "status", "isMaster", "regDate", "approveDate")
    return {k: user.get(k) for k in allowed if k in user}


def _find_auth_user(users: list[dict], email: str) -> tuple[int, dict | None]:
    normalized = email.strip().lower()
    for idx, user in enumerate(users):
        if str(user.get("email", "")).strip().lower() == normalized:
            return idx, user
    return -1, None


def _hash_password(password: str) -> str:
    if PASSWORD_HASH_ALGO in {"argon2", "argon2id"} and _argon2_hasher:
        return "argon2$" + _argon2_hasher.hash(password)
    if PASSWORD_HASH_ALGO == "bcrypt" and bcrypt:
        rounds = int(os.environ.get("SCOREQUERY_BCRYPT_ROUNDS", "12"))
        hashed = bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt(rounds=rounds))
        return "bcrypt$" + hashed.decode("utf-8")
    return "werkzeug$" + generate_password_hash(password)


def _verify_password(stored_hash: str, password: str) -> tuple[bool, bool]:
    if not stored_hash:
        return False, False

    if stored_hash.startswith("argon2$") and _argon2_hasher:
        encoded = stored_hash[len("argon2$") :]
        try:
            ok = _argon2_hasher.verify(encoded, password)
            return ok, bool(ok and _argon2_hasher.check_needs_rehash(encoded))
        except Exception:
            return False, False

    if stored_hash.startswith("bcrypt$") and bcrypt:
        encoded = stored_hash[len("bcrypt$") :].encode("utf-8")
        ok = bcrypt.checkpw(password.encode("utf-8"), encoded)
        return bool(ok), bool(ok and PASSWORD_HASH_ALGO != "bcrypt")

    if stored_hash.startswith("werkzeug$"):
        encoded = stored_hash[len("werkzeug$") :]
        ok = check_password_hash(encoded, password)
        return bool(ok), bool(ok and PASSWORD_HASH_ALGO in {"argon2", "argon2id", "bcrypt"})

    if re.fullmatch(r"[0-9a-f]{64}", stored_hash):
        ok = hmac.compare_digest(hashlib.sha256(password.encode("utf-8")).hexdigest(), stored_hash)
        return bool(ok), bool(ok)

    return False, False


def _validate_password_strength(password: str) -> str | None:
    if len(password) < 10:
        return "비밀번호는 10자 이상이어야 합니다."
    checks = [
        re.search(r"[a-z]", password),
        re.search(r"[A-Z]", password),
        re.search(r"\d", password),
        re.search(r"[^A-Za-z0-9]", password),
    ]
    if sum(1 for c in checks if c) < 3:
        return "비밀번호는 영문 대/소문자, 숫자, 특수문자 중 3종 이상을 포함해야 합니다."
    return None


def _current_auth_user() -> dict | None:
    user = session.get("scorequery_user")
    return user if isinstance(user, dict) else None


def _require_server_master_or_admin() -> tuple[bool, str | None]:
    ok, _ = _require_admin()
    if ok:
        return True, None
    user = _current_auth_user()
    if user and bool(user.get("isMaster")):
        return True, None
    return False, "마스터 세션 또는 관리자 API 토큰이 필요합니다."


def _log_secret() -> str:
    return os.environ.get("SCOREQUERY_LOG_SALT") or app.secret_key


def _student_view_id(course_id: str, sid) -> str:
    msg = f"{course_id}|{sid}".encode("utf-8")
    return hmac.new(_log_secret().encode("utf-8"), msg, hashlib.sha256).hexdigest()


def _hash_for_log(value: str) -> str:
    msg = str(value or "").encode("utf-8")
    return hmac.new(_log_secret().encode("utf-8"), msg, hashlib.sha256).hexdigest()


def _append_view_log(event: dict) -> None:
    os.makedirs(os.path.dirname(VIEW_LOG_FILE) or ".", exist_ok=True)
    line = json.dumps(event, ensure_ascii=False, separators=(",", ":"))
    with _view_log_lock:
        with open(VIEW_LOG_FILE, "a", encoding="utf-8") as f:
            f.write(line + "\n")


def _load_view_logs() -> list[dict]:
    if not os.path.exists(VIEW_LOG_FILE):
        return []
    events: list[dict] = []
    with _view_log_lock:
        with open(VIEW_LOG_FILE, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    event = json.loads(line)
                except Exception:
                    continue
                if isinstance(event, dict):
                    events.append(event)
    return events


app = Flask(__name__, static_folder="docs/static")
app.secret_key = os.environ.get(SECRET_KEY_ENV) or secrets.token_urlsafe(32)
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE=SESSION_COOKIE_SAMESITE,
    SESSION_COOKIE_SECURE=SESSION_COOKIE_SECURE,
)
if TRUST_PROXY:
    app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_port=1)

ENCRYPTED_DATA_FILE = os.path.join("docs", "data.enc.json")
PLAINTEXT_DATA_FILE = os.path.join("docs", "data.json")
PUBLIC_CONFIG_FILE = os.path.join("docs", "public-config.json")

_auth_lock = threading.Lock()
_view_log_lock = threading.Lock()
_argon2_hasher = PasswordHasher() if PasswordHasher else None


# ──────────────────────────────────────────────
# Excel 데이터 로드 및 전처리
# ──────────────────────────────────────────────
EXCEL_FILE = os.environ.get("SCOREQUERY_EXCEL", "2026-1학기_경영정보론_서창갑.xlsx")

students: dict[int, dict] = {}      # {student_id: row_dict}
class_averages: dict[int, dict] = {}
class_maxes: dict[int, dict] = {}
class_counts: dict[int, int] = {}
course_metadata: dict[str, object] = {}
dynamic_eval_meta: list[dict] = []


def _course_id(course: dict) -> str:
    raw = "_".join(str(course.get(k, "") or "") for k in ("year", "semester", "name"))
    value = re.sub(r"\s+", "_", raw)
    value = re.sub(r"[^\w가-힣-]", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "scorequery-course"


def _read_local_config() -> dict:
    if not os.path.exists(CONFIG_FILE):
        return {}
    try:
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _derive_course_metadata() -> dict:
    cfg = _read_local_config()
    configured = cfg.get("course") if isinstance(cfg.get("course"), dict) else {}

    filename = os.path.basename(EXCEL_FILE)
    match = re.match(r"^(\d{4})-(.+?)_(.+?)_(.+)\.xlsx$", filename)
    fallback = {
        "year": match.group(1) if match else "2026",
        "semester": match.group(2) if match else "1학기",
        "name": match.group(3) if match else "성적조회",
    }

    course = {
        "year": str(configured.get("year") or fallback["year"]),
        "semester": str(configured.get("semester") or fallback["semester"]),
        "name": str(configured.get("name") or fallback["name"]),
    }
    course["id"] = str(configured.get("id") or _course_id(course))

    professor_cfg = cfg.get("professor") if isinstance(cfg.get("professor"), dict) else {}
    professor = {
        "name": str(professor_cfg.get("name") or (match.group(4) if match else "")),
        "email": str(professor_cfg.get("email") or cfg.get("professor_email") or ""),
    }
    if professor["name"] or professor["email"]:
        course["professor"] = professor

    course["published"] = bool(students)
    return course


def safe_float(v):
    """셀 값을 float로 안전 변환 (None, 수식문자열 대응)"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


def mask_name(name: str) -> str:
    """이름 마스킹: 첫 글자만 표시, 나머지 '*'"""
    if not name:
        return ""
    return name[0] + "*" * (len(name) - 1)


def mask_student_id(sid) -> str:
    """학번 마스킹: 앞 4자리 표시, 나머지 '****'"""
    s = str(sid)
    if len(s) <= 4:
        return s
    return s[:4] + "*" * (len(s) - 4)


def extract_phone_last4(phone: str) -> str:
    """전화번호에서 숫자만 추출 후 마지막 4자리 반환"""
    if not phone:
        return ""
    digits = re.sub(r"[^0-9]", "", str(phone))
    return digits[-4:] if len(digits) >= 4 else digits


def has_relative_exclusion_marker(value) -> bool:
    text = re.sub(r"\s+", "", str(value or ""))
    return "상대평가제외" in text


def unique_non_empty(values) -> list[str]:
    result = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return result


def find_header_idx(headers, keywords, exclude_keywords=None):
    """헤더 행에서 키워드 기반 컬럼 인덱스 탐색"""
    for idx, h in enumerate(headers):
        if any(kw in h for kw in keywords):
            if exclude_keywords and any(ex in h for ex in exclude_keywords):
                continue
            return idx
    return None


def load_excel():
    """서버 시작 시 Excel 파일을 로드하여 메모리에 캐싱"""
    global students, class_averages, class_maxes, class_counts, course_metadata

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    # 1. 사용할 시트 결정 (최종성적 -> 총괄 -> 첫 번째 시트 순)
    sheet_name = None
    for name in ["최종성적", "총괄"]:
        if name in wb.sheetnames:
            sheet_name = name
            break
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"[ScoreQuery] 시트 '{ws.title}'에서 데이터를 로드합니다.")

    # 2. 헤더 행 읽기 및 키워드 기반 동적 매핑
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in first_row]

    col = {
        "department":   find_header_idx(headers, ["소속", "학과", "학부", "전공"]),
        "class_num":    find_header_idx(headers, ["분반", "반"]),
        "student_id":   find_header_idx(headers, ["학번"]),
        "name":         find_header_idx(headers, ["성명", "이름"]),
        "phone":        find_header_idx(headers, ["전화", "핸드폰", "연락처", "휴대폰"]),
        "exclude":      find_header_idx(headers, ["상대평가제외사유", "상대평가 제외 사유", "상대평가제외", "상대평가 제외", "제외사유", "제외"]),
        "extra":        find_header_idx(headers, ["가산점"]),
        "extra_memo":   find_header_idx(headers, ["가산메모"]),
        "special":      find_header_idx(headers, ["특별점수", "특별"]),
        "special_memo": find_header_idx(headers, ["특별점수메모", "특별메모"]),
        "total":        find_header_idx(headers, ["합계", "총점", "성적"]),
        "rank":         find_header_idx(headers, ["석차", "순위", "등수"], exclude_keywords=["결석"]),
        "grade":        find_header_idx(headers, ["학점", "평점", "등급"]),
        "absences":     find_header_idx(headers, ["결석", "결석횟수", "결석차시"]),
        "remark":       find_header_idx(headers, ["비고"]),
    }

    # 동적 평가항목 추출
    EVAL_MAPPING = {
        "quiz": {"label": "퀴즈", "icon": "🎯", "keywords": ["퀴즈"]},
        "attendance": {"label": "출석", "icon": "📋", "keywords": ["출석"]},
        "assignment": {"label": "과제", "icon": "📝", "keywords": ["과제"]},
        "midterm": {"label": "중간고사", "icon": "📖", "keywords": ["중간"]},
        "final": {"label": "기말고사", "icon": "📕", "keywords": ["기말"]},
        "presentation": {"label": "발표", "icon": "🎤", "keywords": ["발표"]},
        "participation": {"label": "참여도", "icon": "🙋", "keywords": ["참여", "참여도"]},
    }
    
    global dynamic_eval_meta
    dynamic_eval_meta = []
    dynamic_evals = [] # [{'id': 'quiz', 'col_idx': 5}, ...]
    for eval_id, meta in EVAL_MAPPING.items():
        idx = find_header_idx(headers, meta["keywords"])
        if idx is not None:
            dynamic_evals.append({"id": eval_id, "col_idx": idx, "label": meta["label"]})
            dynamic_eval_meta.append({
                "id": eval_id,
                "label": meta["label"],
                "icon": meta["icon"],
                "ratio": 100 # 기본값
            })

    # 필수 컬럼(학번, 이름)이 감지되지 않으면 에러
    if col["student_id"] is None or col["name"] is None:
        raise ValueError("❌ Excel 파일에서 필수 컬럼('학번', '이름')을 찾을 수 없습니다.")

    # 분반별 점수 집계용
    class_scores = {}  # {class_num: {field: [values]}}

    new_students: dict[int, dict] = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) <= col["student_id"]:
            continue

        sid = row[col["student_id"]]
        if sid is None:
            continue

        try:
            # 학번이 숫자가 아닌 경우 건너뛰기
            sid = int(float(str(sid).strip()))
        except (ValueError, TypeError):
            continue

        class_num = 1
        if col["class_num"] is not None and col["class_num"] < len(row) and row[col["class_num"]] is not None:
            try:
                class_num = int(float(str(row[col["class_num"]]).strip()))
            except ValueError:
                class_num = 1

        phone_val = row[col["phone"]] if col["phone"] is not None and col["phone"] < len(row) else ""
        phone_last4 = extract_phone_last4(phone_val)

        def get_val(key, default=None):
            idx = col[key]
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        total = safe_float(get_val("total"))
        rank_val = get_val("rank")
        grade = get_val("grade") or ""
        exclude_reason = get_val("exclude") or ""
        extra = safe_float(get_val("extra"))
        extra_memo = get_val("extra_memo") or ""
        special = safe_float(get_val("special"))
        special_memo = get_val("special_memo") or ""

        absences_val = get_val("absences", 0)
        try:
            absences = int(float(str(absences_val).strip())) if absences_val is not None else 0
        except ValueError:
            absences = 0

        remark = get_val("remark") or ""
        dept = get_val("department") or ""
        student_name = get_val("name") or ""
        relative_exclusion_reason = " / ".join(
            unique_non_empty([
                exclude_reason,
                extra_memo,
                remark if has_relative_exclusion_marker(remark) else "",
            ])
        )
        is_relative_excluded = bool(relative_exclusion_reason)

        # 평가항목 합계 구하기
        eval_sum = 0.0
        for ev in dynamic_evals:
            val = None
            if ev["col_idx"] < len(row):
                val = safe_float(row[ev["col_idx"]])
            if val is not None:
                eval_sum += val

        calculated_total = eval_sum + (extra or 0.0) + (special or 0.0)
        calculated_total = round(calculated_total, 2)

        if total is None or total == 0.0:
            total = calculated_total

        student = {
            "department": dept,
            "class_num": class_num,
            "student_id": sid,
            "name": student_name,
            "phone_last4": phone_last4,
            "extra_score": extra,
            "extra_memo": extra_memo,
            "special_score": special,
            "special_memo": special_memo,
            "total_score": total,
            "rank": rank_val,
            "grade": grade,
            "absences": absences,
            "remark": remark,
            "exclude_reason": exclude_reason,
            "relative_exclusion_reason": relative_exclusion_reason,
            "is_relative_excluded": is_relative_excluded,
        }

        for ev in dynamic_evals:
            val = None
            if ev["col_idx"] < len(row):
                val = safe_float(row[ev["col_idx"]])
            student[ev["id"] + "_score"] = val

        new_students[sid] = student

        # 분반별 점수 집계
        if class_num not in class_scores:
            class_scores[class_num] = { "total_score": [] }
            for ev in dynamic_evals:
                class_scores[class_num][ev["id"] + "_score"] = []
            class_scores[class_num]["extra_score"] = []
            class_scores[class_num]["special_score"] = []

        # 결시자(비고에 '결시' 또는 '미응시' 포함)는 평균/최고점수 집계에서 제외
        is_absent = "결시" in (student["remark"] or "") or "미응시" in (student["remark"] or "")
        if not is_absent:
            for field in class_scores[class_num]:
                val = student.get(field)
                if val is not None:
                    class_scores[class_num][field].append(val)

    # 분반별 석차 자동 계산 (만약 석차가 비어있다면)
    class_groups_app = {}
    for sid, st in new_students.items():
        cn = st["class_num"]
        if cn not in class_groups_app:
            class_groups_app[cn] = []
        class_groups_app[cn].append(st)

    for cn, group in class_groups_app.items():
        for st in group:
            if st["rank"] is None or str(st["rank"]).strip() in ["", "-", "None"]:
                my_total = st["total_score"] or 0.0
                rank = sum(1 for other in group if (other["total_score"] or 0.0) > my_total) + 1
                st["rank"] = str(rank)
            else:
                st["rank"] = str(st["rank"]).strip()

    # 분반별 평균·최고 계산
    new_class_averages, new_class_maxes, new_class_counts = {}, {}, {}
    for cn, scores in class_scores.items():
        avg = {}
        mx = {}
        for field, vals in scores.items():
            if vals:
                avg[field] = round(sum(vals) / len(vals), 2)
                mx[field] = round(max(vals), 2)
            else:
                avg[field] = None
                mx[field] = None
        new_class_averages[cn] = avg
        new_class_maxes[cn] = mx
        new_class_counts[cn] = sum(1 for s in new_students.values() if s["class_num"] == cn)

    # 전체를 한 번에 교체 (부분 로드 상태 방지)
    students = new_students
    class_averages = new_class_averages
    class_maxes = new_class_maxes
    class_counts = new_class_counts
    course_metadata = _derive_course_metadata()

    wb.close()
    print(f"[ScoreQuery] {len(students)}명 학생 데이터 로드 완료 (분반 {len(class_averages)}개)")


# ──────────────────────────────────────────────
# CORS (제한적 화이트리스트)
# ──────────────────────────────────────────────
@app.before_request
def enforce_https_if_configured():
    if not REQUIRE_HTTPS:
        return None
    if request.method == "OPTIONS":
        return None
    if request.is_secure:
        return None
    host = request.host.split(":", 1)[0]
    if host in {"127.0.0.1", "localhost"}:
        return None
    return jsonify({"error": "HTTPS 연결만 허용됩니다."}), 403


@app.after_request
def add_security_and_cors_headers(response):
    response.headers.setdefault("X-Content-Type-Options", "nosniff")
    response.headers.setdefault("Referrer-Policy", "strict-origin-when-cross-origin")
    response.headers.setdefault("Permissions-Policy", "camera=(), microphone=(), geolocation=()")
    if REQUIRE_HTTPS or request.is_secure:
        response.headers.setdefault("Strict-Transport-Security", "max-age=31536000; includeSubDomains")

    origin = request.headers.get("Origin", "")
    if origin and origin in ALLOWED_ORIGINS:
        response.headers["Access-Control-Allow-Origin"] = origin
        response.headers["Vary"] = "Origin"
        response.headers["Access-Control-Allow-Headers"] = "Content-Type, X-Admin-Token"
        response.headers["Access-Control-Allow-Methods"] = "POST, GET, OPTIONS"
        response.headers["Access-Control-Allow-Credentials"] = "true"
    return response


# ──────────────────────────────────────────────
# 관리자 인증 & rate limit 유틸
# ──────────────────────────────────────────────
def _get_admin_token() -> str | None:
    return os.environ.get(ADMIN_TOKEN_ENV) or None


def _require_admin() -> tuple[bool, str | None]:
    """관리자 토큰 검증. (ok, error_message)"""
    expected = _get_admin_token()
    if not expected:
        return False, (
            f"{ADMIN_TOKEN_ENV} 환경변수가 설정되어 있지 않아 관리자 API가 비활성화되었습니다."
        )

    provided = request.headers.get("X-Admin-Token", "")
    if not provided:
        body = request.get_json(silent=True) or {}
        provided = body.get("admin_token", "")

    if not provided or not hmac.compare_digest(str(provided), str(expected)):
        return False, "관리자 토큰이 올바르지 않습니다."

    return True, None


def _client_ip() -> str:
    # X-Forwarded-For 헤더를 최우선으로 확인 (Nginx 등 프록시 환경 대응)
    forwarded = request.headers.get("X-Forwarded-For")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.remote_addr or "unknown"


def _rate_limited(key: str) -> bool:
    now = time.monotonic()
    bucket = _rate_buckets[key]
    while bucket and now - bucket[0] > RATE_WINDOW_SEC:
        bucket.popleft()
    if len(bucket) >= RATE_MAX_PER_WINDOW:
        return True
    bucket.append(now)
    return False


# ──────────────────────────────────────────────
# 관리자 API: 암호화 저장
# ──────────────────────────────────────────────
@app.route("/api/save_data", methods=["POST", "OPTIONS"])
def save_data():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_admin()
    if not ok:
        return jsonify({"error": err}), 401

    try:
        data = request.get_json(silent=True)
        if not data:
            return jsonify({"error": "올바르지 않은 데이터 형식입니다."}), 400

        # admin_token 은 페이로드에서 제거하여 저장
        if isinstance(data, dict) and "admin_token" in data:
            data = {k: v for k, v in data.items() if k != "admin_token"}

        passphrase = get_env_passphrase()
        if not passphrase:
            raise EncryptionConfigError(
                f"{PASSPHRASE_ENV} 환경변수를 설정해야 암호화 저장을 할 수 있습니다."
            )

        write_encrypted_json(ENCRYPTED_DATA_FILE, data, passphrase)
        if os.path.exists(PLAINTEXT_DATA_FILE):
            os.remove(PLAINTEXT_DATA_FILE)

        print(f"[ScoreQuery] encrypted data saved ({os.path.getsize(ENCRYPTED_DATA_FILE)} bytes)")
        return jsonify({
            "success": True,
            "message": "성적 데이터가 AES-256-GCM으로 암호화되어 저장되었습니다.",
            "path": ENCRYPTED_DATA_FILE,
        })
    except EncryptionConfigError as e:
        print(f"[ScoreQuery] encryption configuration error: {e}")
        return jsonify({
            "error": str(e),
            "encryption_required": True,
            "env": PASSPHRASE_ENV,
        }), 500
    except Exception as e:
        print(f"❌ 데이터 저장 중 오류 발생: {e}")
        return jsonify({"error": f"데이터 저장 실패: {str(e)}"}), 500


@app.route("/api/save_public_config", methods=["POST", "OPTIONS"])
def save_public_config():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_admin()
    if not ok:
        return jsonify({"error": err}), 401

    try:
        data = request.get_json(silent=True) or {}
        gas_url = str(data.get("gas_url", "")).strip()
        api_url = str(data.get("api_url", "")).strip().rstrip("/")

        # 매우 가벼운 검증: https URL만 허용
        if gas_url and not gas_url.startswith("https://"):
            return jsonify({"error": "gas_url은 https:// 로 시작해야 합니다."}), 400
        if api_url and not (
            api_url.startswith("https://")
            or api_url.startswith("http://127.0.0.1")
            or api_url.startswith("http://localhost")
        ):
            return jsonify({"error": "api_url은 https:// 또는 로컬 개발 주소만 허용됩니다."}), 400

        config_data = {"gas_url": gas_url, "api_url": api_url}
        os.makedirs(os.path.dirname(PUBLIC_CONFIG_FILE), exist_ok=True)
        with open(PUBLIC_CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config_data, f, ensure_ascii=False, indent=2)

        print(f"[ScoreQuery] public-config.json saved with gas_url: {gas_url}, api_url: {api_url}")
        return jsonify({
            "success": True,
            "message": "공개 연동 설정(public-config.json)이 저장되었습니다.",
            "path": PUBLIC_CONFIG_FILE,
        })
    except Exception as e:
        print(f"❌ 설정 저장 중 오류 발생: {e}")
        return jsonify({"error": f"설정 저장 실패: {str(e)}"}), 500


# ──────────────────────────────────────────────
# 서버 세션 기반 교수 인증 API (운영 백엔드용)
# ──────────────────────────────────────────────
@app.route("/api/auth/bootstrap", methods=["POST", "OPTIONS"])
def auth_bootstrap():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_admin()
    if not ok:
        return jsonify({"error": err}), 401

    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    name = str(data.get("name", "")).strip() or "ScoreQuery Master"

    if not email or not password:
        return jsonify({"error": "email, password가 필요합니다."}), 400
    password_error = _validate_password_strength(password)
    if password_error:
        return jsonify({"error": password_error}), 400

    with _auth_lock:
        users = _load_auth_users()
        if users and not data.get("force"):
            return jsonify({"error": "이미 서버 인증 사용자가 있습니다."}), 409
        idx, existing = _find_auth_user(users, email)
        user = {
            "email": email,
            "name": name,
            "univ": str(data.get("univ", "")).strip(),
            "dept": str(data.get("dept", "")).strip(),
            "phone": str(data.get("phone", "")).strip(),
            "pw": _hash_password(password),
            "status": "approved",
            "isMaster": True,
            "regDate": existing.get("regDate") if existing else _now_iso(),
            "approveDate": _now_iso(),
        }
        if idx >= 0:
            users[idx] = user
        else:
            users.append(user)
        _save_auth_users(users)

    return jsonify({"success": True, "user": _sanitize_auth_user(user)})


@app.route("/api/auth/register", methods=["POST", "OPTIONS"])
def auth_register():
    if request.method == "OPTIONS":
        return "", 204

    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    if not email or not password:
        return jsonify({"error": "email, password가 필요합니다."}), 400
    password_error = _validate_password_strength(password)
    if password_error:
        return jsonify({"error": password_error}), 400

    with _auth_lock:
        users = _load_auth_users()
        if not users:
            return jsonify({"error": "최초 마스터는 /api/auth/bootstrap으로 생성해야 합니다."}), 409
        idx, existing = _find_auth_user(users, email)
        if existing and existing.get("status") != "rejected":
            return jsonify({"error": "이미 등록된 이메일입니다."}), 409

        user = {
            "email": email,
            "name": str(data.get("name", "")).strip(),
            "univ": str(data.get("univ", "")).strip(),
            "dept": str(data.get("dept", "")).strip(),
            "phone": str(data.get("phone", "")).strip(),
            "pw": _hash_password(password),
            "status": "pending",
            "isMaster": False,
            "regDate": _now_iso(),
            "approveDate": "",
        }
        if idx >= 0:
            users[idx] = user
        else:
            users.append(user)
        _save_auth_users(users)

    return jsonify({"success": True, "user": _sanitize_auth_user(user)})


@app.route("/api/auth/login", methods=["POST", "OPTIONS"])
def auth_login():
    if request.method == "OPTIONS":
        return "", 204

    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    if not email or not password:
        return jsonify({"error": "email, password가 필요합니다."}), 400

    with _auth_lock:
        users = _load_auth_users()
        if not users:
            return jsonify({"error": "최초 마스터는 /api/auth/bootstrap으로 생성해야 합니다."}), 409
        idx, user = _find_auth_user(users, email)
        ok, needs_rehash = _verify_password(str(user.get("pw", "")) if user else "", password)
        if not user or not ok:
            return jsonify({"error": "이메일 또는 비밀번호가 올바르지 않습니다."}), 401
        if user.get("status") != "approved":
            return jsonify({"error": "승인된 계정만 로그인할 수 있습니다.", "status": user.get("status")}), 403
        if needs_rehash:
            user["pw"] = _hash_password(password)
            users[idx] = user
            _save_auth_users(users)

    safe_user = _sanitize_auth_user(user)
    session["scorequery_user"] = safe_user
    session.permanent = False
    return jsonify({"success": True, "user": safe_user})


@app.route("/api/auth/me", methods=["GET"])
def auth_me():
    user = _current_auth_user()
    return jsonify({"authenticated": bool(user), "user": user})


@app.route("/api/auth/logout", methods=["POST", "OPTIONS"])
def auth_logout():
    if request.method == "OPTIONS":
        return "", 204
    session.pop("scorequery_user", None)
    return jsonify({"success": True})


@app.route("/api/auth/change_password", methods=["POST", "OPTIONS"])
def auth_change_password():
    if request.method == "OPTIONS":
        return "", 204

    current = _current_auth_user()
    if not current:
        return jsonify({"error": "로그인이 필요합니다."}), 401

    data = request.get_json(silent=True) or {}
    old_password = str(data.get("old_password", ""))
    new_password = str(data.get("new_password", ""))
    password_error = _validate_password_strength(new_password)
    if password_error:
        return jsonify({"error": password_error}), 400

    with _auth_lock:
        users = _load_auth_users()
        idx, user = _find_auth_user(users, str(current.get("email", "")))
        ok, _ = _verify_password(str(user.get("pw", "")) if user else "", old_password)
        if not user or not ok:
            return jsonify({"error": "현재 비밀번호가 올바르지 않습니다."}), 401
        user["pw"] = _hash_password(new_password)
        users[idx] = user
        _save_auth_users(users)

    return jsonify({"success": True})


@app.route("/api/auth/users", methods=["GET", "OPTIONS"])
def auth_users():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_server_master_or_admin()
    if not ok:
        return jsonify({"error": err}), 401
    users = [_sanitize_auth_user(user) for user in _load_auth_users()]
    return jsonify({"success": True, "users": users})


@app.route("/api/auth/set_status", methods=["POST", "OPTIONS"])
def auth_set_status():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_server_master_or_admin()
    if not ok:
        return jsonify({"error": err}), 401

    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    status = str(data.get("status", "")).strip()
    if status not in {"pending", "approved", "rejected", "deleted"}:
        return jsonify({"error": "허용되지 않는 status입니다."}), 400

    with _auth_lock:
        users = _load_auth_users()
        idx, user = _find_auth_user(users, email)
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
        user["status"] = status
        if status == "approved":
            user["approveDate"] = _now_iso()
        users[idx] = user
        _save_auth_users(users)

    return jsonify({"success": True, "user": _sanitize_auth_user(user)})


# ──────────────────────────────────────────────
# Routes
# ──────────────────────────────────────────────
@app.route("/api/auth/reset_password", methods=["POST", "OPTIONS"])
def auth_reset_password():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_server_master_or_admin()
    if not ok:
        return jsonify({"error": err}), 401

    data = request.get_json(silent=True) or {}
    email = str(data.get("email", "")).strip().lower()
    password = str(data.get("password", ""))
    password_error = _validate_password_strength(password)
    if not email or password_error:
        return jsonify({"error": password_error or "email, password가 필요합니다."}), 400

    with _auth_lock:
        users = _load_auth_users()
        idx, user = _find_auth_user(users, email)
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
        user["pw"] = _hash_password(password)
        users[idx] = user
        _save_auth_users(users)

    return jsonify({"success": True, "user": _sanitize_auth_user(user)})


@app.route("/api/auth/withdraw", methods=["POST", "OPTIONS"])
def auth_withdraw():
    if request.method == "OPTIONS":
        return "", 204

    current = _current_auth_user()
    if not current:
        return jsonify({"error": "로그인이 필요합니다."}), 401
    if current.get("isMaster"):
        return jsonify({"error": "마스터 계정은 탈퇴할 수 없습니다."}), 400

    data = request.get_json(silent=True) or {}
    current_password = str(data.get("current_password", ""))

    with _auth_lock:
        users = _load_auth_users()
        idx, user = _find_auth_user(users, str(current.get("email", "")))
        if not user:
            return jsonify({"error": "사용자를 찾을 수 없습니다."}), 404
        if current_password:
            ok, _ = _verify_password(str(user.get("pw", "")), current_password)
            if not ok:
                return jsonify({"error": "현재 비밀번호가 올바르지 않습니다."}), 401
        user["status"] = "deleted"
        user["withdrawReqDate"] = _now_iso()
        user["withdrawApproveDate"] = _now_iso()
        users[idx] = user
        _save_auth_users(users)

    session.pop("scorequery_user", None)
    return jsonify({"success": True, "user": _sanitize_auth_user(user)})


@app.route("/")
def index():
    """메인 페이지 서빙 (docs/index.html을 직접 서비스)"""
    return send_from_directory("docs", "index.html")


@app.route("/docs/<path:filename>")
def serve_docs(filename):
    return send_from_directory("docs", filename)


@app.route("/docs/")
@app.route("/docs")
def serve_docs_index():
    return send_from_directory("docs", "index.html")


@app.route("/admin")
def admin_redirect():
    return redirect("/docs/")


@app.route("/api/courses", methods=["GET"])
def get_courses():
    metadata = course_metadata or _derive_course_metadata()
    if not students:
        return jsonify({"courses": []})
    return jsonify({"courses": [metadata]})


def _record_score_view(course: dict, sid: int, student: dict) -> None:
    course_id = str(course.get("id") or _course_id(course))
    event = {
        "event": "score_view",
        "timestamp": _now_iso(),
        "course_id": course_id,
        "student_view_id": _student_view_id(course_id, sid),
        "class_num": student.get("class_num"),
        "ip_hash": _hash_for_log(_client_ip()),
        "user_agent_hash": _hash_for_log(request.headers.get("User-Agent", "")),
    }
    try:
        _append_view_log(event)
    except Exception as exc:
        print(f"[ScoreQuery] view log write failed: {exc}")


@app.route("/api/view_stats", methods=["GET", "OPTIONS"])
def get_view_stats():
    if request.method == "OPTIONS":
        return "", 204

    ok, err = _require_admin()
    if not ok:
        return jsonify({"error": err}), 401

    metadata = course_metadata or _derive_course_metadata()
    course_id = str(request.args.get("course_id") or metadata.get("id") or _course_id(metadata))
    events = [
        event for event in _load_view_logs()
        if event.get("event") == "score_view" and str(event.get("course_id")) == course_id
    ]

    latest_by_student: dict[str, str] = {}
    for event in events:
        view_id = str(event.get("student_view_id") or "")
        timestamp = str(event.get("timestamp") or "")
        if not view_id or not timestamp:
            continue
        existing = latest_by_student.get(view_id)
        if not existing or timestamp > existing:
            latest_by_student[view_id] = timestamp

    student_rows = []
    for sid, student in students.items():
        view_id = _student_view_id(course_id, sid)
        view_date = latest_by_student.get(view_id)
        student_rows.append({
            "student_id_masked": mask_student_id(sid),
            "name_masked": mask_name(student.get("name", "")),
            "department": student.get("department", ""),
            "class_num": student.get("class_num"),
            "is_viewed": bool(view_date),
            "view_date": view_date,
        })

    viewed_count = sum(1 for row in student_rows if row["is_viewed"])
    total_count = len(student_rows)
    return jsonify({
        "success": True,
        "course": metadata,
        "course_id": course_id,
        "total": total_count,
        "viewed": viewed_count,
        "unviewed": max(total_count - viewed_count, 0),
        "percent": round((viewed_count / total_count) * 100, 1) if total_count else 0,
        "students": student_rows,
    })


@app.route("/api/score", methods=["POST"])
def get_score():
    """
    학생 성적 조회 API
    요청: { "student_id": "20220034", "phone_last4": "5169", "access_code": "123456" }
    """
    # 1) Rate limit (IP 기준)
    if _rate_limited(_client_ip()):
        return jsonify({
            "error": "잠시 후 다시 시도해 주세요. (요청 횟수 제한)",
        }), 429

    # 2) 입력 검증
    data = request.get_json(silent=True)
    if not data or not isinstance(data, dict):
        return jsonify({"error": "요청 데이터가 올바르지 않습니다."}), 400

    raw_id = str(data.get("student_id", "")).strip()
    phone_last4 = str(data.get("phone_last4", "")).strip()
    access_code = str(data.get("access_code", "")).strip()

    if not raw_id or not phone_last4 or not access_code:
        return jsonify({"error": "학번, 전화번호 뒷자리, 접속 비밀번호를 모두 입력해 주세요."}), 400

    if len(raw_id) > MAX_STUDENT_ID_LEN or not raw_id.isdigit():
        return jsonify({"error": "학번은 숫자만 입력 가능합니다."}), 400

    if not phone_last4.isdigit() or len(phone_last4) != 4:
        return jsonify({"error": "전화번호 뒷자리 4자리를 정확히 입력해 주세요."}), 400

    if not access_code.isdigit() or len(access_code) != 6:
        return jsonify({"error": "접속 비밀번호 6자리 숫자를 정확히 입력해 주세요."}), 400

    # 3) 접속 비밀번호(access_code) 설정 값과 대조 검증 (설정되어 있는 경우에만 검증)
    expected_code = _get_configured_access_code()
    if not expected_code:
        return jsonify({
            "error": "성적 조회 접속 비밀번호가 서버에 설정되어 있지 않습니다. 관리자에게 문의해 주세요."
        }), 503
    if not hmac.compare_digest(access_code, expected_code):
        return jsonify({
            "error": "일치하는 정보를 찾을 수 없습니다.\n비밀번호를 다시 확인해 주세요."
        }), 404

    sid = int(raw_id)
    student = students.get(sid)

    # 4) 학번/전화번호 검증을 한 묶음으로 처리 (계정 열거 방지)
    if not student or not hmac.compare_digest(student["phone_last4"], phone_last4):
        return jsonify({
            "error": "일치하는 정보를 찾을 수 없습니다.\n학번과 전화번호를 다시 확인해 주세요."
        }), 404

    # 분반 정보
    cn = student["class_num"]
    avg = class_averages.get(cn, {})
    mx = class_maxes.get(cn, {})
    count = class_counts.get(cn, 0)

    # 석차 포맷팅 제거 -> 엑셀 원본 값 그대로 사용
    rank_val = student["rank"]
    rank_str = str(rank_val) if rank_val is not None else "-"

    student_res = {
        "department": student["department"],
        "class_num": cn,
        "student_id_masked": mask_student_id(sid),
        "name_masked": mask_name(student["name"]),
        "total_score": student["total_score"],
        "rank": rank_str,
        "grade": student["grade"],
        "absences": student["absences"],
        "remark": student["remark"],
        "extra_memo": student.get("extra_memo", ""),
        "special_memo": student.get("special_memo", ""),
        "relative_exclusion_reason": student.get("relative_exclusion_reason", ""),
        "is_relative_excluded": bool(student.get("is_relative_excluded")),
    }
    
    # 동적 평가항목 점수 추가
    for k in student.keys():
        if k.endswith("_score") and k != "total_score":
            student_res[k] = student[k]

    course = course_metadata or _derive_course_metadata()
    _record_score_view(course, sid, student)

    response = {
        "student": student_res,
        "course": course,
        "evaluation": dynamic_eval_meta,
        "class_avg": avg,
        "class_max": mx,
        "class_count": count,
    }

    return jsonify(response)


# ──────────────────────────────────────────────
# Entry Point
# ──────────────────────────────────────────────
if __name__ == "__main__":
    try:
        load_excel()
    except Exception as e:
        print(f"[Warning] Excel file load failed: {e}")
        print("Starting server without pre-loaded Excel data.")

    if not _get_admin_token():
        print(
            f"[ScoreQuery] ⚠️  {ADMIN_TOKEN_ENV} 환경변수가 설정되지 않아 "
            "관리자 API(/api/save_data, /api/save_public_config, /api/view_stats)는 비활성화됩니다."
        )

    if not os.environ.get(SECRET_KEY_ENV):
        print(
            f"[ScoreQuery] ⚠️  {SECRET_KEY_ENV} 환경변수가 없어 서버 세션 키가 재시작마다 바뀝니다. "
            "운영 환경에서는 반드시 긴 임의 값을 설정하세요."
        )

    if DEFAULT_HOST not in ("127.0.0.1", "localhost"):
        print(
            f"[ScoreQuery] ⚠️  외부 접근 가능한 호스트({DEFAULT_HOST})로 바인딩합니다. "
            "개발 서버 직접 공개 대신 WSGI 서버 + HTTPS 리버스 프록시를 사용하세요."
        )

    if REQUIRE_HTTPS and not TRUST_PROXY:
        print("[ScoreQuery] ⚠️  HTTPS 강제 모드입니다. 프록시 뒤 운영이면 SCOREQUERY_TRUST_PROXY=1 설정을 확인하세요.")

    app.run(host=DEFAULT_HOST, port=DEFAULT_PORT, debug=False)
