# -*- coding: utf-8 -*-
"""
build_data.py — Excel → 정적 JSON 변환 스크립트
GitHub Pages 배포용 데이터 파일을 생성합니다.

개인정보 보호:
  - 키: SHA-256(학번 + "|" + 전화번호뒷4자리) → 원본 역추적 불가
  - 전화번호, 이메일: JSON에 포함하지 않음
  - 이름: 마스킹 처리 (첫 글자만 표시)
  - 학번: 앞 4자리만 표시

사용법:
  python build_data.py
  → docs/data.enc.json 생성
"""

import hashlib
import json
import os
import re

import openpyxl

from scorequery_crypto import (
    get_env_passphrase,
    get_passphrase,
    load_encrypted_json,
    write_encrypted_json,
)

EXCEL_FILE = os.environ.get("SCOREQUERY_EXCEL", "2026-1학기_경영정보론_서창갑.xlsx")
PLAINTEXT_OUTPUT_FILE = "docs/data.json"
OUTPUT_FILE = "docs/data.enc.json"
ENCRYPTED_CONFIG_FILE = "config.enc.json"


def safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    try:
        return round(float(v), 2)
    except (ValueError, TypeError):
        return None


def mask_name(name):
    if not name:
        return ""
    return name[0] + "*" * (len(name) - 1)


def mask_student_id(sid):
    s = str(sid)
    return s[:4] + "*" * (len(s) - 4) if len(s) > 4 else s


def extract_phone_last4(phone):
    if not phone:
        return ""
    digits = re.sub(r"[^0-9]", "", str(phone))
    return digits[-4:] if len(digits) >= 4 else digits


def make_hash_key(student_id, phone_last4, access_code):
    """SHA-256 해시 키 생성 - 원본 역추적 방지"""
    raw = f"{student_id}|{phone_last4}|{access_code}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def has_relative_exclusion_marker(value):
    text = str(value or "").replace(" ", "")
    return "상대평가제외" in text


def unique_non_empty(values):
    result = []
    seen = set()
    for value in values:
        text = str(value or "").strip()
        if text and text not in seen:
            result.append(text)
            seen.add(text)
    return result


def build():
    # 🔑 접속 비밀번호 6자리 획득
    access_code = os.environ.get("SCOREQUERY_ACCESS_CODE", "")
    if not access_code and os.path.exists("config.json"):
        try:
            with open("config.json", "r", encoding="utf-8") as f:
                import json
                cfg = json.load(f)
                access_code = str(cfg.get("access_code", "")).strip()
        except Exception:
            pass

    if not access_code:
        while True:
            val = input("학생 조회에 사용할 6자리 접속 비밀번호를 설정하세요 (예: 123456): ").strip()
            if val.isdigit() and len(val) == 6:
                access_code = val
                break
            print("⚠️ 6자리 숫자로 정확히 입력해 주세요.")

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    
    # 1. 사용할 시트 결정 (최종성적 -> 총괄 -> 첫 번째 시트 순)
    sheet_name = None
    for name in ["최종성적", "총괄"]:
        if name in wb.sheetnames:
            sheet_name = name
            break
    ws = wb[sheet_name] if sheet_name else wb.active
    print(f"[build_data] 시트 '{ws.title}'에서 데이터를 읽어옵니다.")

    # 2. 헤더 행 읽기 및 키워드 기반 동적 매핑
    first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
    
    def find_idx(keywords, exclude_keywords=None):
        for idx, h in enumerate(headers):
            if any(kw in h for kw in keywords):
                if exclude_keywords and any(ex in h for ex in exclude_keywords):
                    continue
                return idx
        return None

    col = {
        "department":   find_idx(["소속", "학과", "학부", "전공"]),
        "class_num":    find_idx(["분반", "반"]),
        "student_id":   find_idx(["학번"]),
        "name":         find_idx(["성명", "이름"]),
        "phone":        find_idx(["전화", "핸드폰", "연락처", "휴대폰"]),
        "extra":        find_idx(["가산점"]),
        "extra_memo":   find_idx(["가산메모"]),
        "special":      find_idx(["특별점수", "특별"]),
        "special_memo": find_idx(["특별점수메모", "특별메모"]),
        "total":        find_idx(["합계", "총점", "성적"]),
        "rank":         find_idx(["석차", "순위", "등수"], exclude_keywords=["결석"]),
        "grade":        find_idx(["학점", "평점", "등급"]),
        "absences":     find_idx(["결석", "결석횟수", "결석차시"]),
        "remark":       find_idx(["비고"]),
    }

    # 동적 평가항목 추출
    EVAL_MAPPING = {
        "quiz": {"label": "퀴즈", "icon": "🎯", "keywords": ["퀴즈"]},
        "attendance": {"label": "출석", "icon": "📋", "keywords": ["출석"]},
        "assignment": {"label": "과제", "icon": "📝", "keywords": ["과제"]},
        "midterm": {"label": "중간고사", "icon": "📖", "keywords": ["중간"]},
        "final": {"label": "기말고사", "icon": "📕", "keywords": ["기말"]},
        "presentation": {"label": "발표", "icon": "🎤", "keywords": ["발표"]},
        "participation": {"label": "참여도", "icon": "🙋", "keywords": ["참여", "참여도"]},
    }
    
    dynamic_eval_meta = []
    dynamic_evals = []
    for eval_id, meta in EVAL_MAPPING.items():
        idx = find_idx(meta["keywords"])
        if idx is not None:
            dynamic_evals.append({"id": eval_id, "col_idx": idx, "label": meta["label"]})
            dynamic_eval_meta.append({
                "id": eval_id,
                "label": meta["label"],
                "icon": meta["icon"],
                "ratio": 100
            })

    # 필수 컬럼(학번, 이름)이 감지되지 않으면 에러
    if col["student_id"] is None or col["name"] is None:
        raise ValueError("❌ Excel 파일에서 필수 컬럼('학번', '이름')을 찾을 수 없습니다.")

    students = {}
    class_scores = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) <= col["student_id"]:
            continue
            
        sid = row[col["student_id"]]
        if sid is None:
            continue

        try:
            # 학번이 숫자가 아닌 경우 건너뛰기
            sid = int(float(str(sid).strip()))
        except (ValueError, TypeError):
            continue

        class_num = 1
        if col["class_num"] is not None and col["class_num"] < len(row) and row[col["class_num"]] is not None:
            try:
                class_num = int(float(str(row[col["class_num"]]).strip()))
            except ValueError:
                class_num = 1

        phone_val = row[col["phone"]] if col["phone"] is not None and col["phone"] < len(row) else ""
        phone_last4 = extract_phone_last4(phone_val)

        # 해시 키 생성
        hash_key = make_hash_key(sid, phone_last4, access_code)

        def get_val(key, default=None):
            idx = col[key]
            if idx is not None and idx < len(row):
                return row[idx]
            return default

        total = safe_float(get_val("total"))
        rank_val = get_val("rank")
        grade = get_val("grade") or ""
        extra = safe_float(get_val("extra"))
        extra_memo = get_val("extra_memo") or ""
        special = safe_float(get_val("special"))
        special_memo = get_val("special_memo") or ""
        
        absences_val = get_val("absences", 0)
        try:
            absences = int(float(str(absences_val).strip())) if absences_val is not None else 0
        except ValueError:
            absences = 0
            
        remark = get_val("remark") or ""
        dept = get_val("department") or ""
        student_name = get_val("name") or ""
        relative_exclusion_reason = " / ".join(
            unique_non_empty([
                extra_memo,
                remark if has_relative_exclusion_marker(remark) else "",
            ])
        )
        is_relative_excluded = bool(relative_exclusion_reason)

        # 평가항목 합계 구하기
        eval_sum = 0.0
        for ev in dynamic_evals:
            val = None
            if ev["col_idx"] < len(row):
                val = safe_float(row[ev["col_idx"]])
            if val is not None:
                eval_sum += val

        calculated_total = eval_sum + (extra or 0.0) + (special or 0.0)
        calculated_total = round(calculated_total, 2)

        if total is None or total == 0.0:
            total = calculated_total

        student_data = {
            "department": dept,
            "class_num": class_num,
            "student_id_masked": mask_student_id(sid),
            "name_masked": mask_name(student_name),
            "extra_score": extra,
            "extra_memo": extra_memo,
            "special_score": special,
            "special_memo": special_memo,
            "total_score": total,
            "rank": rank_val,
            "grade": grade,
            "absences": absences,
            "remark": remark,
            "relative_exclusion_reason": relative_exclusion_reason,
            "is_relative_excluded": is_relative_excluded,
        }
        
        for ev in dynamic_evals:
            val = None
            if ev["col_idx"] < len(row):
                val = safe_float(row[ev["col_idx"]])
            student_data[ev["id"] + "_score"] = val
            
        students[hash_key] = student_data

        # 분반별 집계
        if class_num not in class_scores:
            class_scores[class_num] = {
                "total_score": [],
                "count": 0,
            }
            for ev in dynamic_evals:
                class_scores[class_num][ev["id"] + "_score"] = []
            class_scores[class_num]["extra_score"] = []
            class_scores[class_num]["special_score"] = []
                
        class_scores[class_num]["count"] += 1
        
        # 결시자(비고에 '결시' 또는 '미응시' 포함)는 평균/최고점수 집계에서 제외
        is_absent = "결시" in remark or "미응시" in remark
        if not is_absent:
            # 동적 평가항목 및 가산점/특별점수 추가
            fields_to_aggregate = ["total_score"] + [ev["id"] + "_score" for ev in dynamic_evals] + ["extra_score", "special_score"]
            for field in fields_to_aggregate:
                val = students[hash_key].get(field)
                if val is not None:
                    class_scores[class_num][field].append(val)

    # 분반별 석차 자동 계산 (만약 석차가 비어있다면)
    class_groups_py = {}
    for hk, st in students.items():
        cn = st["class_num"]
        if cn not in class_groups_py:
            class_groups_py[cn] = []
        class_groups_py[cn].append(st)

    for cn, group in class_groups_py.items():
        for st in group:
            if st["rank"] is None or str(st["rank"]).strip() in ["", "-", "None"]:
                my_total = st["total_score"] or 0.0
                rank = sum(1 for other in group if (other["total_score"] or 0.0) > my_total) + 1
                st["rank"] = str(rank)
            else:
                st["rank"] = str(st["rank"]).strip()

    # 분반별 평균·최고 계산
    class_averages = {}
    class_maxes = {}
    class_counts = {}

    for cn, data in class_scores.items():
        avg = {}
        mx = {}
        fields_to_aggregate = ["total_score"] + [ev["id"] + "_score" for ev in dynamic_evals]
        for field in fields_to_aggregate:
            vals = data[field]
            if vals:
                avg[field] = round(sum(vals) / len(vals), 2)
                mx[field] = round(max(vals), 2)
            else:
                avg[field] = None
                mx[field] = None
        class_averages[str(cn)] = avg
        class_maxes[str(cn)] = mx
        class_counts[str(cn)] = data["count"]
    # 기존 암호화 파일이나 로컬 설정에서 gas_url을 읽어오기
    gas_url = ""
    env_passphrase = get_env_passphrase()

    try:
        if env_passphrase and os.path.exists(OUTPUT_FILE):
            old_data = load_encrypted_json(OUTPUT_FILE, env_passphrase)
            gas_url = old_data.get("gas_url", "")
    except Exception:
        pass

    try:
        if env_passphrase and os.path.exists(ENCRYPTED_CONFIG_FILE):
            cfg = load_encrypted_json(ENCRYPTED_CONFIG_FILE, env_passphrase)
            if cfg.get("gas_url"):
                gas_url = cfg.get("gas_url")
    except Exception:
        pass

    try:
        if os.path.exists("config.json"):
            with open("config.json", "r", encoding="utf-8") as f:
                cfg = json.load(f)
                if cfg.get("gas_url"):
                    gas_url = cfg.get("gas_url")
    except Exception:
        pass

    # 파일명 기반 메타데이터 추출 로직
    filename = os.path.basename(EXCEL_FILE)
    m = re.match(r'^(\d{4})-(.+?)_(.+?)_(.+)\.xlsx$', filename)
    if m:
        course_year = m.group(1)
        course_semester = m.group(2)
        course_name = m.group(3)
        prof_name = m.group(4)
    else:
        course_year = "2026"
        course_semester = "1학기"
        course_name = "경영정보시스템"
        prof_name = "서창갑"

    # JSON 출력
    output = {
        "course": {
            "year": course_year,
            "semester": course_semester,
            "name": course_name
        },
        "professor": {
            "name": prof_name,
            "email": "armour@tu.ac.kr"
        },
        "evaluation": dynamic_eval_meta,
        "access_code": access_code,
        "gas_url": gas_url,
        "students": students,
        "class_avg": class_averages,
        "class_max": class_maxes,
        "class_counts": class_counts,
    }
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)

    passphrase = get_passphrase(confirm=not bool(env_passphrase))
    write_encrypted_json(OUTPUT_FILE, output, passphrase)
    if os.path.exists(PLAINTEXT_OUTPUT_FILE):
        os.remove(PLAINTEXT_OUTPUT_FILE)

    # public-config.json 생성 (가입/로그인 시 GAS URL 조회를 위한 공개 설정 파일)
    public_config_path = os.path.join("docs", "public-config.json")
    try:
        with open(public_config_path, "w", encoding="utf-8") as f:
            json.dump({"gas_url": gas_url}, f, ensure_ascii=False, indent=2)
        print(f"[build_data] {public_config_path} 생성 완료 (gas_url: {gas_url})")
    except Exception as e:
        print(f"⚠️ [build_data] public-config.json 생성 실패: {e}")

    wb.close()
    print(f"[build_data] {len(students)}명 데이터 -> {OUTPUT_FILE} 암호화 생성 완료")
    print(f"   분반: {len(class_averages)}개, 파일크기: {os.path.getsize(OUTPUT_FILE):,} bytes")
    print("   보안방식: AES-256-GCM + PBKDF2-HMAC-SHA256, SHA-256(학번|전화번호뒷4자리)")


if __name__ == "__main__":
    build()
